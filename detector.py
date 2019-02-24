import cv2
import cv2.cv as cv
import numpy as np
import openpyxl

recognizer = cv2.createLBPHFaceRecognizer()
recognizer.load('trainer2/trainer.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);


cam = cv2.VideoCapture(0)
font = cv2.cv.InitFont(cv2.cv.CV_FONT_HERSHEY_SIMPLEX, 1, 1, 0, 1, 1)
while True:
    ret, im =cam.read()
    if ret is True:
        gray=cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)
    else:
        continue
    t = cv.GetTickCount()    
    faces=faceCascade.detectMultiScale(gray, 1.2,5)
    t = cv.GetTickCount() - t
    #print "detection time = %gms" % (t/(cv.GetTickFrequency()*1000.))
    for(x,y,w,h) in faces:
        cv2.rectangle(im,(x,y),(x+w,y+h),(225,0,0),2)
        Id, conf = recognizer.predict(gray[y:y+h,x:x+w])
        if(conf<50):
            if(Id==1):
                Id="Arthi"
            elif(Id==2):
                Id="Surez"
            
        else:
            Id="Unknown"
        cv2.cv.PutText(cv2.cv.fromarray(im),str(Id), (x,y+h),font, 255)
    cv2.imshow('im',im) 
    if cv2.waitKey(10) & 0xFF==ord('q'):
        break
print "detection time = %gms" % (t/(cv.GetTickFrequency()*1000.))
print("Updating..............")
wb = openpyxl.load_workbook(filename = 'attendance.xlsx')
for ws in wb.worksheets:
    print(ws.title)
         
ws = wb.worksheets[0]
    
a=["Names","Arthi","Surez","Nivi","Akil","Naresh"]
b=["Attendance","Present"]
c=["Id",1,2]

for i in range(len(a)):
    ws.cell(row=i+1, column=1).value = a[i]

for j in range(len(b)):
    ws.cell(row=1,column=2).value=b[0]
    if(Id==a[j]):
        print(Id,a[j])
        ws.cell(row=j+1,column=2).value="Present"
    else:
        ws.cell(row=j+1,column=2).value="Absent"
        
wb.save("attendance.xlsx")
print("Attendance Sheet Updated......................")
#print ("Accuracy=",faces.score(gray,faces))
cam.release()
cv2.destroyAllWindows()
